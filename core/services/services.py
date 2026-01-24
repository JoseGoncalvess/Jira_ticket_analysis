import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import os
import re


# def unificar_planilhas(pasta_origem, nome_arquivo_final, log_callback):
#     log_callback(f"\n--- Iniciando a unificação das planilhas ---")
    
#     # 1. Verifica se a pasta existe
#     if not os.path.exists(pasta_origem):
#         log_callback(f"[ERRO] A pasta {pasta_origem} não foi encontrada.")
#         return

#     # 2. Lista todos os arquivos Excel na pasta
#     arquivos = [f for f in os.listdir(pasta_origem) if f.endswith('.xlsx')]
    
#     # Remove o arquivo final da lista caso ele já exista na mesma pasta (para não duplicar)
#     nome_final_sem_caminho = os.path.basename(nome_arquivo_final)
#     if nome_final_sem_caminho in arquivos:
#         arquivos.remove(nome_final_sem_caminho)

#     if not arquivos:
#         log_callback("[AVISO] Nenhum arquivo .xlsx encontrado para unificar.")
#         return

#     lista_dfs = []
    
#     # 3. Loop para ler cada arquivo
#     for arquivo in arquivos:
#         caminho_completo = os.path.join(pasta_origem, arquivo)
#         try:
#             # Lê o Excel
#             df_temp = pd.read_excel(caminho_completo)
            
#             # Opcional: Adicionar uma coluna para saber de qual arquivo veio (útil para auditoria)
#             # df_temp['Arquivo_Origem'] = arquivo 
            
#             lista_dfs.append(df_temp)
#             log_callback(f"  -> Lido: {arquivo} ({len(df_temp)} linhas)")
#         except Exception as e:
#             log_callback(f"  [ERRO] Falha ao ler {arquivo}: {e}")

#     # 4. Junta tudo e Salva
#     if lista_dfs:
#         log_callback("Consolidando dados...")
#         df_consolidado = pd.concat(lista_dfs, ignore_index=True)
        
#         try:
#             df_consolidado.to_excel(nome_arquivo_final, index=False)
#             log_callback(f"SUCESSO! Relatório Unificado gerado em:\n{nome_arquivo_final}")
#             log_callback(f"Total de linhas consolidadas: {len(df_consolidado)}")
#         except Exception as e:
#             log_callback(f"[ERRO] Não foi possível salvar o arquivo consolidado: {e}")
#     else:
#         log_callback("[ERRO] A lista de dados está vazia.")


def unificar_planilhas(pasta_origem, nome_arquivo_final, log_callback):
    log_callback(f"\n--- Iniciando a unificação das planilhas (Modo Leve) ---")
    
    # 1. Verifica se a pasta existe
    if not os.path.exists(pasta_origem):
        log_callback(f"[ERRO] A pasta {pasta_origem} não foi encontrada.")
        return

    # 2. Lista todos os arquivos Excel na pasta
    arquivos = [f for f in os.listdir(pasta_origem) if f.endswith('.xlsx')]
    
    # Remove o arquivo final da lista caso ele já exista (para não duplicar infinitamente)
    nome_final_sem_caminho = os.path.basename(nome_arquivo_final)
    if nome_final_sem_caminho in arquivos:
        arquivos.remove(nome_final_sem_caminho)

    if not arquivos:
        log_callback("[AVISO] Nenhum arquivo .xlsx encontrado para unificar.")
        return

    # --- PREPARAÇÃO DO ARQUIVO MESTRE ---
    wb_master = openpyxl.Workbook()
    ws_master = wb_master.active
    ws_master.title = "RELATORIO_GERAL_CONSOLIDADO"

    
    
    cabecalho_adicionado = False
    total_linhas_consolidadas = 0

    # 3. Loop para ler cada arquivo
    for arquivo in arquivos:
        caminho_completo = os.path.join(pasta_origem, arquivo)
        try:
            # Carrega o arquivo temporário (data_only=True pega o valor, não a fórmula)
            wb_temp = openpyxl.load_workbook(caminho_completo, data_only=True)
            ws_temp = wb_temp.active
            
            # Pega todas as linhas como valores
            # values_only=True retorna tuplas com os dados limpos: ('Valor A', 'Valor B')
            linhas = list(ws_temp.iter_rows(values_only=True))
            
            if not linhas:
                log_callback(f"  -> Ignorado (Vazio): {arquivo}")
                wb_temp.close()
                continue


            if not cabecalho_adicionado:
                ws_master.append(linhas[0]) # Adiciona cabeçalho
                
                # --- APLICA NEGRITO NO MESTRE ---
                for celula in ws_master[1]:
                    celula.font = Font(bold=True)
                # --------------------------------
                
                cabecalho_adicionado = True
                dados_novos = linhas[1:]

            # Se for o primeiro arquivo a ser processado, pegamos TUDO (incluindo cabeçalho)
            if not cabecalho_adicionado:
                # Adiciona o cabeçalho (linha 0)
                ws_master.append(linhas[0])
                cabecalho_adicionado = True
                
                # Adiciona o restante dos dados (da linha 1 até o fim)
                dados_novos = linhas[1:]
            else:
                # Se NÃO é o primeiro, pulamos o cabeçalho e pegamos só os dados
                dados_novos = linhas[1:]

            

            # Escreve as linhas no mestre
            for linha in dados_novos:
                ws_master.append(linha)
                total_linhas_consolidadas += 1

            log_callback(f"  -> Lido: {arquivo} ({len(dados_novos)} novos registros)")
            
            # Fecha para liberar memória
            wb_temp.close()

        except Exception as e:
            log_callback(f"  [ERRO] Falha ao ler {arquivo}: {e}")
    

   
    # 4. Salva o arquivo final
    if total_linhas_consolidadas > 0:
        try:
            log_callback("Salvando arquivo consolidado...")
            
            wb_master.save(nome_arquivo_final)
            log_callback(f"SUCESSO! Relatório Unificado gerado em:\n{nome_arquivo_final}")
            log_callback(f"Total de linhas consolidadas: {total_linhas_consolidadas}")
        except Exception as e:
            log_callback(f"[ERRO] Não foi possível salvar o arquivo consolidado: {e}")
    else:
        log_callback("[ERRO] Nenhum dado foi encontrado para consolidar.")

def convert_to_date(date_string) -> str:
    if not isinstance(date_string, str):
        return date_string
    try:
        formato_original = "%a, %d %b %Y %H:%M:%S %z"
        objeto_data = datetime.strptime(date_string, formato_original)
        formato_desejado = "%d/%m/%Y"
        return objeto_data.strftime(formato_desejado)
    except (ValueError, TypeError):
        return date_string  
    

def limpar_cto(valor_celula):
    # Verificação de segurança: se a célula estiver vazia ou não for texto, retorna como está
    if not isinstance(valor_celula, str):
        return valor_celula

    # 1. Separa por vírgula
    itens = valor_celula.split(',')

    # 2. Filtra: Mantém o item SE "CTO" NÃO estiver nele (usando lower() para pegar cto, CTO, Cto)
    # O .strip() remove espaços extras que possam ter sobrado
    itens_filtrados = [item.strip() for item in itens if "cto" not in item.lower()]

    # 3. Junta de volta com vírgula
    return ",".join(itens_filtrados)



# def processar_arquivo_xml(caminho_do_arquivo, dados_agregados, log_callback):
   
#     file = os.path.basename(caminho_do_arquivo)
#     log_callback(f"---> Processando arquivo: {file}")
#     try:
#         tree = ET.parse(caminho_do_arquivo)
#         root = tree.getroot()
#     except ET.ParseError as e:
#         log_callback(f"     [ERRO] Arquivo XML mal formatado: {e}")
#         return

#     lista_de_itens = root.findall('.//item')
    
  
#     chamados_neste_arquivo = []
#     for item in lista_de_itens:
#         titulo_node = item.find('title')
#         data_crated_node = item.find('created')
#         link_node = item.find('link')
#         key_node = item.find('key')
#         status_node = item.find('status')
#         response_node = item.find('assignee')
        
#         titulo = titulo_node.text.strip() if titulo_node is not None and titulo_node.text else "Título não encontrado"
#         dataCreated = data_crated_node.text if data_crated_node is not None else ""
#         link = link_node.text if link_node is not None else ""
#         key = key_node.text if key_node is not None else "Key não encontrado"
#         status = status_node.text if status_node is not None else "Status não encontrado"
#         responsavel = response_node.text if response_node is not None else "Responsavel não encontrado não encontrado"

#         chamados_vinculados = []
#         cto_value=''
#         issuelinks_node = item.find('issuelinks')
#         if issuelinks_node is not None:
#             for link_node in issuelinks_node.findall('.//issuekey'):
#                 if link_node.text.__contains__("CTO"):
#                         cto_value = link_node.text
#                 if link_node.text:
#                     chamados_vinculados.append(link_node.text)
#         # chamados_neste_arquivo.append({
#         #     "Tipo":key.split("-")[0],
#         #     'Titulo': titulo,
#         #     'Vinculados': chamados_vinculados,
#         #     'Criado em': convert_to_date(dataCreated),
#         #     'Link': link
#         # })

#         chamados_neste_arquivo.append({
#             'Criação': convert_to_date(dataCreated),
#             'status':status,
#             "Chave":key,
#             "Alteração de Status":"",
#             'Vinculado (AVB/MOB/FLUX)':", ".join(chamados_vinculados),
#             "Responsável":responsavel,
#             "CTO":cto_value,
#             'Link': link
#         })



#     log_callback(f"     Encontrados {len(chamados_neste_arquivo)} chamados no arquivo {file}.")
    
    
#     for tiket in chamados_neste_arquivo:
#         # list_of_vinc = tiket["Vinculados"]
#         list_of_vinc = tiket["Vinculado (AVB/MOB/FLUX)"]
        
#         for key in dados_agregados.keys():
#             if key in list_of_vinc:
#                 dados_agregados[key].append(tiket)



def processar_arquivo_xml(caminho_do_arquivo, dados_agregados, log_callback):
    file = os.path.basename(caminho_do_arquivo)
    log_callback(f"---> Processando arquivo: {file}")
    
    try:
        tree = ET.parse(caminho_do_arquivo)
        root = tree.getroot()
    except ET.ParseError as e:
        log_callback(f"     [ERRO] Arquivo XML mal formatado: {e}")
        return

    lista_de_itens = root.findall('.//item')
    count_processados = 0

    for item in lista_de_itens:
        # 1. Extração dos Dados Básicos
        titulo_node = item.find('title')
        titulo = titulo_node.text.strip() if titulo_node is not None and titulo_node.text else ""
        
        # ... (Extração de dataCreated, link, key, status, responsavel permanece igual) ...
        # Resumido para focar na lógica:
        dataCreated = item.find('created').text if item.find('created') is not None else ""
        link = item.find('link').text if item.find('link') is not None else ""
        key = item.find('key').text if item.find('key') is not None else ""
        status = item.find('status').text if item.find('status') is not None else ""
        responsavel = item.find('assignee').text if item.find('assignee') is not None else ""

        # 2. Identificação dos Donos (CTOs)
        chaves_encontradas = set() # Usamos set para não repetir o mesmo CTO no mesmo chamado
        chamados_vinculados_texto = [] 

        # A) Procura nos Links (issuelinks)
        issuelinks_node = item.find('issuelinks')
        if issuelinks_node is not None:
            for link_node in issuelinks_node.findall('.//issuekey'):
                texto_link = link_node.text
                if texto_link:
                    chamados_vinculados_texto.append(texto_link)
                    if "CTO" in texto_link:
                        chaves_encontradas.add(texto_link)

        # B) Procura no Título (Regex)
        ctos_no_titulo = re.findall(r'(CTO-\d+)', titulo, re.IGNORECASE)
        for cto in ctos_no_titulo:
            chaves_encontradas.add(cto.upper())

        # Define o valor da coluna "CTO" (Vazio se não tiver dono)
        coluna_cto_valor = next(iter(chaves_encontradas)) if chaves_encontradas else ""

        # Monta o objeto do chamado
        ticket_dict = {
            'Criação': convert_to_date(dataCreated),
            'status': status,
            "Chave": key,
            "Alteração de Status": "",
            'Vinculado (AVB/MOB/FLUX)': ", ".join(chamados_vinculados_texto),
            "Responsável": responsavel,
            "CTO": coluna_cto_valor, # Vai ficar vazio nos casos "SEM-CTO"
            'Link': link
        }
        
        count_processados += 1

        # ---------------------------------------------------------
        # AQUI ESTÁ A NOVA LÓGICA DE DISTRIBUIÇÃO (O PULO DO GATO)
        # ---------------------------------------------------------
        
        # Cenário 1: O chamado NÃO TEM nenhum CTO identificado
        if not chaves_encontradas:
            chave_destino = "SEM-CTO"
            
            # Verifica se a gaveta "SEM-CTO" existe no seu dicionário pré-carregado
            # Se não existir, a gente cria na hora para garantir que não dê erro
            if chave_destino not in dados_agregados:
                dados_agregados[chave_destino] = []
            
            # Adiciona o chamado na lista "SEM-CTO"
            dados_agregados[chave_destino].append(ticket_dict)

        # Cenário 2: O chamado TEM donos (CTO-XXX)
        else:
            for cto_key in chaves_encontradas:
                # Verificamos se esse CTO está na sua lista de interesse (dados_agregados)
                if cto_key in dados_agregados:
                    dados_agregados[cto_key].append(ticket_dict)
                else:
                    # Opcional: Se aparecer um CTO novo que não estava na lista original,
                    # você quer criar uma aba pra ele ou jogar no SEM-CTO?
                    # Por padrão, vou criar a aba pra ele:
                    dados_agregados[cto_key] = [ticket_dict]

    log_callback(f"     Processados {count_processados} itens neste arquivo.")


def criar_planilhas_por_empresa(dados,log_callback,caminho_saida):
   
    log_callback("\n--- Iniciando a criação das planilhas finais ---")
 
    pasta_relatorios = caminho_saida if caminho_saida else "Relatorios_Jira" 
    
    pasta_relatorios = f"{pasta_relatorios}//Relatorios_Jira"

    if not os.path.exists(pasta_relatorios):
        os.makedirs(pasta_relatorios)
        log_callback(f"Pasta '{pasta_relatorios}' criada.")

    for id_empresa, lista_de_chamados in dados.items():
        if not lista_de_chamados:
            log_callback(f"- ID {id_empresa}: Não possui chamados. Pulando...")
            continue
            
        log_callback(f"- ID {id_empresa}: Encontrados {len(lista_de_chamados)} chamados. Gerando planilha...")

        # OLD VERSION - CRIADNO PLANILHA COM PANDAS
        # df = pd.DataFrame(lista_de_chamados)
        
        # df['Vinculado (AVB/MOB/FLUX)'] = df['Vinculado (AVB/MOB/FLUX)'].apply(limpar_cto)
    
        # NEW VERISON

        # Cria um Workbok novo - CRIANDO PLANILHA COM OPENPYXL
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório"

        # Pega os cabeçalhos do primeiro item (chaves do dicionário)
        cabecalhos = list(lista_de_chamados[0].keys())
        ws.append(cabecalhos) # Escreve a linha 1 (cabeçalho)

# --- APLICAÇÃO DO NEGRITO ---
        # Percorre todas as células da linha 1 (cabeçalho)
        for celula in ws[1]:
            celula.font = Font(bold=True)
        # Escreve os dados
        # 3. Itera sobre os dados para escrever as linhas
        for chamado in lista_de_chamados:
            linha_para_escrever = []
            
            for col in cabecalhos:
                valor_original = chamado.get(col)
                
                # --- AQUI É ONDE A MÁGICA ACONTECE (O equivalente ao .apply) ---
                if col == 'Vinculado (AVB/MOB/FLUX)':
                    # Aplica a função de limpeza antes de adicionar na lista
                    valor_tratado = limpar_cto(valor_original)
                    linha_para_escrever.append(valor_tratado)
                else:
                    # Se não for a coluna de vinculados, adiciona o valor normal
                    linha_para_escrever.append(valor_original)
            
            # Adiciona a linha completa na planilha
            ws.append(linha_para_escrever)


        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Pega a letra da coluna (A, B, C...)
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            
            adjusted_width = (max_length + 3)
            ws.column_dimensions[column].width = adjusted_width

        nome_arquivo = f"Relatorio_{id_empresa}.xlsx"
        caminho_completo = os.path.join(pasta_relatorios, nome_arquivo)

        # df.to_excel(caminho_completo, index=False)
        wb.save(caminho_completo)


        log_callback(f"  -> Planilha '{caminho_completo}' criada com sucesso!")

    log_callback("\nProcesso finalizado.")